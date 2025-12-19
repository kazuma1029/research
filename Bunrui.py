import pandas as pd
import torch
from sklearn.metrics import accuracy_score, precision_recall_fscore_support
from transformers import BertJapaneseTokenizer, BertForSequenceClassification
from torch.utils.data import DataLoader, Dataset
import os
import openpyxl


if torch.backends.mps.is_available():
    device = torch.device("mps")
elif torch.cuda.is_available():
    device = torch.device("cuda")
else:
    device = torch.device("cpu")

print(f"Using device: {device}")
def load_finetuned_model(model_path):
    model = BertForSequenceClassification.from_pretrained(model_path)
    tokenizer = BertJapaneseTokenizer.from_pretrained(model_path)
    return model, tokenizer

class ReviewDataset(Dataset):
    def __init__(self, reviews, labels, tokenizer, max_length=512):
        self.reviews = reviews
        self.labels = labels
        self.tokenizer = tokenizer
        self.max_length = max_length

    def __len__(self):
        return len(self.reviews)

    def __getitem__(self, idx):
        review = self.reviews[idx]
        label = self.labels[idx]

        encodings = self.tokenizer(
            review,
            truncation=True,
            padding="max_length",
            max_length=self.max_length,
            return_tensors="pt"
        )
        encodings["labels"] = torch.tensor(label, dtype=torch.long)
        return encodings

def preprocess_data(movie_ids_file, base_path, hikensya_num):
    reviews = []
    labels = []
    all_ratings = []

    with open(movie_ids_file, 'r') as file:
        movie_ids = [line.strip() for line in file.readlines()]

    for movie_id in movie_ids:
        file_path = os.path.join(base_path, f"Experiment_movieID_{movie_id}.xlsx")
        if not os.path.exists(file_path):
            print(f"File not found: {file_path}")
            continue

        data = pd.read_excel(file_path)
        file_reviews = data.iloc[1:, 0].dropna().tolist()
        ratings = data.iloc[1:, 3].dropna().astype(float).tolist()

        all_ratings.extend(ratings)
        reviews.extend(file_reviews)

    if not all_ratings:
        raise ValueError("No ratings found. Please check the input files.")

    threshold = sum(all_ratings) / len(all_ratings)
    labels = [1 if rating >= threshold else 0 for rating in all_ratings]
    num_pos = labels.count(1)
    num_neg = labels.count(0)

    print(f"全映画の評価値平均: {threshold:.2f}")
    print(f"レビュー数: {len(reviews)}, ラベル数: {len(labels)}")
    print(f"正解ラベル(1)の数: {num_pos}, 不正解ラベル(0)の数: {num_neg}")
    return reviews, labels

def compute_metrics(predictions, true_labels):
    preds = torch.argmax(predictions, axis=1).numpy()
    labels = true_labels.numpy()

    accuracy = accuracy_score(labels, preds)
    precision, recall, f1, _ = precision_recall_fscore_support(labels, preds, average="binary")

    return {
        "accuracy": accuracy,
        "precision": precision,
        "recall": recall,
        "f1": f1
    }

def save_results_to_excel(hikensya_num, metrics,ran):
    # ファイルパス
    file_path = f"./実験結果/被験者{hikensya_num}.xlsx"

    # ファイルを開く（存在しない場合は新規作成）
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "結果"
        workbook.save(file_path)

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # 適合率、再現率、F値を記録
    precision = metrics["precision"]
    recall = metrics["recall"]
    f1 = metrics["f1"]

    col=10 #5(E)列目に記録

    # 適合率（3～12行目)
    # for row in range(3, 9):
    row=ran+2
    sheet.cell(row=row, column=col).value = precision

    # 再現率 (11〜16行目)
    # for row in range(11, 17):
    row=ran+13
    sheet.cell(row=row, column=col).value = recall

    # F値 (19〜24行目)
    # for row in range(19, 25):
    row=ran+25
    sheet.cell(row=row, column=col).value = f1

    # 保存
    workbook.save(file_path)
    print(f"結果を {file_path} に保存しました。")

def main(hikensya_num):
    for ran in range(2,4):
        model_path = f"./models/{hikensya_num}_allmodel_rus_{ran}"
        print(f"被験者{hikensya_num}のランダムモデル:model{ran}で分類開始します")
        model, tokenizer = load_finetuned_model(model_path)

        # データ処理と分類
        movie_ids_file = f"./movies/{hikensya_num}_liked_movies.txt"
        base_path = f"./被験者{hikensya_num}/既知の映画群"
        reviews, labels = preprocess_data(movie_ids_file, base_path, hikensya_num)

        dataset = ReviewDataset(reviews, labels, tokenizer)
        dataloader = DataLoader(dataset, batch_size=8)

        all_predictions = []
        all_labels = []

        with torch.no_grad():
            total_reviews = len(dataset)  # 全体のレビュー数を取得
            for idx, batch in enumerate(dataloader):
                # 現在のレビュー番号を計算
                start_idx = idx * dataloader.batch_size
                end_idx = min((idx + 1) * dataloader.batch_size, total_reviews)

                # 現在の進捗を表示
                print(f"レビュー分類中: {start_idx + 1}〜{end_idx} / {total_reviews}")
                
                # 現在のレビュー内容を表示
                current_reviews = reviews[start_idx:end_idx]
                for i, review in enumerate(current_reviews, start=start_idx + 1):
                    print(f"[{i}/{total_reviews}] レビュー内容: {review}")


                # バッチ処理
                inputs = {key: val.squeeze(1).to(torch.device("cpu")) for key, val in batch.items() if key != "labels"}
                outputs = model(**inputs)
                logits = outputs.logits

                all_predictions.append(logits)
                all_labels.append(batch["labels"])

        all_predictions = torch.cat(all_predictions)
        all_labels = torch.cat(all_labels)

        metrics = compute_metrics(all_predictions, all_labels)
        print("分類結果:", metrics)

        # 結果をExcelに保存
        save_results_to_excel(hikensya_num, metrics, ran)

if __name__ == "__main__":
    hikensya_num = input("被験者番号を入力してください：")
    main(hikensya_num)
